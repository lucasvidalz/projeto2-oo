import openpyxl
import pathlib
from fpdf import FPDF
from tkinter import messagebox
import pygame  # Importando pygame
from reports.generate_report import ReportGenerator

class PDFReportGenerator(ReportGenerator):
    def generate_report(self):
        try:
            ficheiro = pathlib.Path("Clientes.xlsx")

            if not ficheiro.exists():
                messagebox.showerror("Erro", "É necessário cadastrar ao menos 1 usuário")
                return

            workbook = openpyxl.load_workbook('Clientes.xlsx')
            folha = workbook.active

            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('Arial', 'B', 12)

            # Header
            pdf.cell(200, 10, 'Relatório de Clientes', ln=True, align='C')

            # Table headers
            pdf.cell(80, 10, 'Nome Completo', border=1)
            pdf.cell(60, 10, 'Contato', border=1)
            pdf.cell(50, 10, 'Idade', border=1)
            pdf.ln()

            for row in folha.iter_rows(min_row=2, values_only=True):
                pdf.cell(80, 10, row[0], border=1)
                pdf.cell(60, 10, row[1], border=1)
                pdf.cell(50, 10, str(row[2]), border=1)
                pdf.ln()

            pdf.output('Relatório_de_Clientes.pdf')
            # Reproduzindo o áudio de sucesso
            pygame.mixer.init()
            pygame.mixer.music.load('res/sounds/relatoriopdf.mp3')
            pygame.mixer.music.play()
            messagebox.showinfo("Sistema", "Relatório gerado com sucesso!")
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro ao gerar o relatório: {e}")
